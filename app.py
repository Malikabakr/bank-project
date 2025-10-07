import os
import logging
import uuid
import zipfile
import time
import threading
from io import BytesIO
from datetime import datetime, timedelta
from flask import (
    Flask,
    render_template,
    request,
    redirect,
    url_for,
    flash,
    session,
    send_file,
    jsonify,
)
import pandas as pd
from werkzeug.utils import secure_filename
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Spacer
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
from arabic_reshaper import reshape
from bidi.algorithm import get_display
import openpyxl
from openpyxl.utils import get_column_letter
from fpdf import FPDF
import re
import fitz
from apscheduler.schedulers.background import BackgroundScheduler

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[logging.FileHandler("file_cleanup.log"), logging.StreamHandler()],
)
logger = logging.getLogger(__name__)

# Define project folder and paths
PROJECT_FOLDER = os.path.dirname(os.path.abspath(__file__))
UPLOAD_FOLDER = os.path.join(PROJECT_FOLDER, "static", "uploads")
GENERATED_PDFS_FOLDER = os.path.join(PROJECT_FOLDER, "static", "uploads", "pdf")
TEMPLATES_FOLDER = os.path.join(PROJECT_FOLDER, "static", "uploads", "templates")
TEMP_FOLDER = os.path.join(PROJECT_FOLDER, "static", "uploads", "temp")

# File cleanup settings
FILE_AGE_LIMIT = 2  # minutes
CLEANUP_INTERVAL = 1  # minutes


def is_file_old_enough(file_path):
    """Check if file is older than FILE_AGE_LIMIT minutes"""
    try:
        mtime = os.path.getmtime(file_path)
        file_age = datetime.now() - datetime.fromtimestamp(mtime)
        return file_age > timedelta(minutes=FILE_AGE_LIMIT)
    except Exception:
        return True


def perform_cleanup():
    """Clean up old files from upload directories"""
    try:
        # List of directories to clean
        cleanup_dirs = [UPLOAD_FOLDER, GENERATED_PDFS_FOLDER, TEMP_FOLDER]

        deleted_files = []
        skipped_files = []
        failed_files = []

        logger.info("Starting scheduled cleanup check...")

        for directory in cleanup_dirs:
            if not os.path.exists(directory):
                logger.info(f"Directory does not exist, skipping: {directory}")
                continue

            # Clean files in directory and subdirectories
            for root, dirs, files in os.walk(directory, topdown=False):
                # Clean files first
                for name in files:
                    file_path = os.path.join(root, name)
                    try:
                        if is_file_old_enough(file_path):
                            # Get file info before deletion
                            file_age = datetime.now() - datetime.fromtimestamp(
                                os.path.getmtime(file_path)
                            )
                            file_size = os.path.getsize(file_path)

                            # Delete the file
                            os.unlink(file_path)

                            # Log deletion with details
                            logger.info(
                                f"Deleted file: {file_path}\n"
                                f"    Age: {int(file_age.total_seconds() / 60)} minutes {int(file_age.total_seconds() % 60)} seconds\n"
                                f"    Size: {file_size / 1024:.2f} KB"
                            )
                            deleted_files.append(file_path)
                        else:
                            age = datetime.now() - datetime.fromtimestamp(
                                os.path.getmtime(file_path)
                            )
                            logger.info(
                                f"Skipped file (too young): {file_path}\n"
                                f"    Current age: {int(age.total_seconds() / 60)} minutes {int(age.total_seconds() % 60)} seconds"
                            )
                            skipped_files.append(file_path)
                    except Exception as e:
                        logger.error(f"Error deleting file {file_path}: {str(e)}")
                        failed_files.append(file_path)

        # Log summary with details of skipped files
        logger.info(
            f"Cleanup Summary:\n"
            f"    Deleted: {len(deleted_files)} items\n"
            f"    Skipped: {len(skipped_files)} items\n"
            f"    Failed: {len(failed_files)} items\n\n"
            f"Skipped Files List:\n" + "\n".join(f"    - {f}" for f in skipped_files)
        )

        # Ensure required directories exist
        for dir_path in [
            UPLOAD_FOLDER,
            GENERATED_PDFS_FOLDER,
            TEMPLATES_FOLDER,
            TEMP_FOLDER,
        ]:
            if not os.path.exists(dir_path):
                os.makedirs(dir_path)
                logger.info(f"Recreated directory: {dir_path}")

    except Exception as e:
        logger.error(f"Error during cleanup: {str(e)}")


# Initialize the scheduler
scheduler = BackgroundScheduler()
scheduler.add_job(
    func=perform_cleanup,
    trigger="interval",
    minutes=CLEANUP_INTERVAL,
    id="cleanup_job",
    name="Clean up old files",
    replace_existing=True,
)
scheduler.start()

# Create Flask app
app = Flask(__name__)
app.secret_key = os.environ.get("SESSION_SECRET", "dev_secret_key")

# Configuration
ALLOWED_EXTENSIONS = {
    "excel": {"xlsx", "xls"},
    "pdf": {"pdf"},
}

# Max file upload size (16MB)
app.config["MAX_CONTENT_LENGTH"] = 16 * 1024 * 1024
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER
app.config["GENERATED_PDFS_FOLDER"] = GENERATED_PDFS_FOLDER
app.config["TEMPLATES_FOLDER"] = TEMPLATES_FOLDER

# Create folders if they don't exist
for folder in [UPLOAD_FOLDER, TEMPLATES_FOLDER, GENERATED_PDFS_FOLDER, TEMP_FOLDER]:
    os.makedirs(folder, exist_ok=True)

# Global progress tracker (thread-safe)
progress_tracker = {}


def update_progress(session_id, value):
    """Update progress value for a session"""
    progress_tracker[session_id] = value


def allowed_file(filename, file_type):
    """Check if a file has an allowed extension"""
    return "." in filename and filename.rsplit(".", 1)[
        1
    ].lower() in ALLOWED_EXTENSIONS.get(file_type, {})


def format_arabic_text(text):
    """Format Arabic text for proper display"""
    if not text:
        return ""
    try:
        # Convert to string and handle any non-string inputs
        text_str = str(text).strip()

        # Check if the text contains Arabic characters
        if any("\u0600" <= char <= "\u06ff" for char in text_str):
            # Reshape and apply bidi
            reshaped_text = reshape(text_str)
            return get_display(reshaped_text)
        return text_str
    except Exception as e:
        logger.error(f"Error formatting text: {e}")
        # Return original as string if there was an error
        if isinstance(text, str):
            return text
        return str(text) if text is not None else ""


@app.route("/")
def index():
    """Home page"""
    # Assign a unique session ID if not already set
    if "session_id" not in session:
        session["session_id"] = str(uuid.uuid4())
    return render_template("index.html")


@app.route("/card_selection")
def card_selection():
    """Card selection page"""
    # Assign a unique session ID if not already set
    if "session_id" not in session:
        session["session_id"] = str(uuid.uuid4())
    return render_template("card_form_selection.html")


@app.route("/platinum_form")
def platinum_form():
    """Platinum card form page"""
    return render_template("platinum_form.html")


@app.route("/business_form")
def business_form():
    """Business card form page"""
    return render_template("business_form.html")


@app.route("/corporate_form")
def corporate_form():
    """Corporate card form page"""
    return render_template("platinum_form.html", card_type="corporate")


@app.route("/isic_form")
def isic_form():
    """ISIC card form page"""
    return render_template("platinum_form.html", card_type="isic")


@app.route("/itic_form")
def itic_form():
    """ITIC card form page"""
    return render_template("platinum_form.html", card_type="itic")


@app.route("/iytc_form")
def iytc_form():
    """IYTC card form page"""
    return render_template("platinum_form.html", card_type="iytc")


@app.route("/a4_form")
def a4_form():
    """A4 form page"""
    return render_template("a4_form.html")


@app.route("/progress")
def get_progress():
    """Get the progress of current conversion job"""
    session_id = session.get("session_id", None)
    if session_id and session_id in progress_tracker:
        progress = progress_tracker[session_id]
        if progress == -1:
            return jsonify(
                {
                    "progress": -1,
                    "error": "An error occurred during processing. Please try again.",
                }
            )
        return jsonify({"progress": progress})
    return jsonify({"progress": 0})


# =====================================
# Excel to PDF Conversion
# =====================================
@app.route("/upload_excel", methods=["POST"])
def upload_excel():
    """Upload Excel file and process according to card type"""
    # Check if file part exists
    if "file" not in request.files:
        return jsonify({"status": "error", "message": "No file selected"})

    file = request.files["file"]

    # Check if filename is empty
    if file.filename == "":
        return jsonify({"status": "error", "message": "No file selected"})

    # Check if file is allowed
    if file and allowed_file(file.filename, "excel"):
        # Create a unique session ID if not already set
        session_id = session.get("session_id", str(uuid.uuid4()))
        session["session_id"] = session_id

        # Get the card type from the form
        card_type = request.form.get(
            "card_type", "platinum"
        )  # Default to 'platinum' if not provided
        session["card_type"] = card_type

        # Check if a PDF template was uploaded (for A4 form)
        pdf_template = None
        if (
            "pdf_template" in request.files
            and request.files["pdf_template"].filename != ""
        ):
            template_file = request.files["pdf_template"]
            template_filename = template_file.filename
            if template_filename and allowed_file(template_filename, "pdf"):
                template_filename = secure_filename(template_filename)
                template_path = os.path.join(
                    TEMPLATES_FOLDER, f"{session_id}_{card_type}_template.pdf"
                )
                template_file.save(template_path)
                pdf_template = template_path
                session["pdf_template"] = template_path

        # Save the Excel file
        filename = file.filename
        excel_path = None
        if filename:
            secure_name = secure_filename(filename)
            excel_path = os.path.join(UPLOAD_FOLDER, f"{session_id}_{secure_name}")
            file.save(excel_path)
            session["excel_file"] = excel_path
        else:
            # Should never happen since we already checked for empty filename above
            # But providing a fallback for safety
            excel_path = os.path.join(UPLOAD_FOLDER, f"{session_id}_data.xlsx")
            file.save(excel_path)
            session["excel_file"] = excel_path

        try:
            # Start processing Excel data in a new thread
            update_progress(session_id, 0)  # Initialize progress

            # Create the output zip filename
            if filename:
                zip_name = os.path.splitext(filename)[0]
                zip_filename = f"{zip_name}_{card_type}.zip"
                session["zip_filename"] = zip_filename
            else:
                zip_filename = f"statements_{card_type}_{session_id}.zip"
                session["zip_filename"] = zip_filename

            # Start processing in the background
            threading.Thread(
                target=process_excel_with_template,
                args=(excel_path, session_id, card_type, pdf_template, zip_filename),
            ).start()

            return jsonify(
                {
                    "status": "started",
                    "message": "PDF generation started.",
                    "zip_filename": zip_filename,
                }
            )

        except Exception as e:
            logger.error(f"Error starting Excel processing: {str(e)}")
            return jsonify({"status": "error", "message": f"Error: {str(e)}"})
    else:
        return jsonify(
            {
                "status": "error",
                "message": "Invalid file type. Please upload an Excel file (.xlsx, .xls).",
            }
        )


# B5 size in points (width, height)
B5_SIZE = (498.9, 708.7)  # Standard B5 size in points

POSITIONS = {
    "platinum": [
        (288, 375),  # activation code
        (258, 406),  # last four digits
        (183, 435),  # name
        (245, 465),  # phone number
        (230, 507),  # delivery address
        (123, 568),  # address title
        (123, 585),  # address description
    ],
    "corporate": [
        (155, 375),  # activation code
        (152, 407),  # name
        (152, 470),  # last four digits
        (152, 525),  # phone number
        (152, 585),  # delivery address
    ],
    "business": [
        (280, 375),  # activation code
        (290, 405),  # last four digits
        (210, 430),  # name
        (185, 445),  # onboarding name
        (240, 470),  # phone number
        (230, 507),  # delivery address
        (123, 568),  # address title
        (123, 585),  # address description
    ],
    "isic": [
        (150, 423),  # last four digits
        (150, 463),  # delivery address
        (150, 503),  # phone number
        (150, 538),  # name
        (150, 575),  # university
    ],
    "itic": [
        (150, 423),  # last four digits
        (150, 463),  # delivery address
        (150, 503),  # phone number
        (150, 538),  # name
        (150, 575),  # university
    ],
    "iytc": [
        (150, 463),  # last four digits
        (150, 503),  # delivery address
        (150, 538),  # phone number
        (150, 575),  # name
    ],
    "a4": [
        (195, 340),  # name
        (195, 380),  # phone number
        (195, 415),  # last four digits
        (195, 450),  # delivery address
    ],
}

FIELD_ORDERS = {
    "platinum": [
        "activation code",
        "last four digits",
        "name",
        "phone number",
        "delivery address",
        "address title",
        "address description",
    ],
    "corporate": [
        "activation code",
        "name",
        "last four digits",
        "phone number",
        "delivery address",
    ],
    "business": [
        "activation code",
        "last four digits",
        "name",
        "onboarding name",
        "phone number",
        "delivery address",
        "address title",
        "address description",
    ],
    "isic": [
        "last four digits",
        "delivery address",
        "phone number",
        "name",
        "university",
    ],
    "itic": [
        "last four digits",
        "delivery address",
        "phone number",
        "name",
        "university",
    ],
    "iytc": ["last four digits", "delivery address", "phone number", "name"],
    "a4": ["name", "phone number", "last four digits", "delivery address"],
}

# 1. Make sure all keys in HEADER_TO_FIELD and FIELD_ORDERS are lowercase
HEADER_TO_FIELD = {
    "card phone number": "phone number",
    "phone number": "phone number",
    "card last digits": "last four digits",
    "last four digits": "last four digits",
    "cardholder name": "name",
    "name": "name",
    "activation code": "activation code",
    "delivery location": "delivery address",
    "delivery address": "delivery address",
    "address title": "address title",
    "address description": "address description",
    "onboarding name": "onboarding name",
    "university": "university",
}


def is_arabic(text):
    text = str(text)
    # Remove invisible characters
    text = re.sub(r"[\u200e\u200f\u202a-\u202e]", "", text)
    # Count Arabic letters
    arabic_count = sum("\u0600" <= char <= "\u06ff" for char in text)
    # If more than half the letters are Arabic, treat as Arabic
    return arabic_count > (len(text.strip()) / 2)


def replace_dashes_in_text(text):
    """Replace dashes in text with appropriate values"""
    if not text:
        return text

    # Convert to string if not already
    text = str(text)

    # Replace common dash patterns
    replacements = {
        "---": "",  # Remove triple dashes
        "--": "",  # Remove double dashes
        "-": "",  # Remove single dashes
    }

    for dash, replacement in replacements.items():
        text = text.replace(dash, replacement)

    return text.strip()


def replace_dashes_in_pdf(template_path, output_path, values):
    """Replace dashes in template PDF with values from Excel, sorted by position."""
    try:
        logger.info(f"[REPLACE_DASHES] Template path: {template_path}")
        logger.info(f"[REPLACE_DASHES] Output path: {output_path}")
        doc = fitz.open(template_path)
        new_doc = fitz.open()
        for page_num in range(len(doc)):
            page = doc[page_num]
            new_page = new_doc.new_page(width=page.rect.width, height=page.rect.height)
            new_page.show_pdf_page(new_page.rect, doc, page_num)
            # Find dashes and sort by (y, x) position
            text_instances = page.search_for("-")
            logger.info(
                f"[REPLACE_DASHES] Page {page_num}: Found {len(text_instances)} dashes."
            )
            text_instances = sorted(text_instances, key=lambda r: (r.y0, r.x0))
            for i, rect in enumerate(text_instances):
                if i >= len(values):
                    logger.info(
                        f"[REPLACE_DASHES] No more values to insert at dash {i}."
                    )
                    break
                new_page.draw_rect(rect, color=(1, 1, 1), fill=(1, 1, 1))
                value = str(values[i])
                logger.info(
                    f"[REPLACE_DASHES] Replacing dash {i} at {rect} with value: {value}"
                )
                # Fix: Use format_arabic_text for Arabic/Kurdish text
                if any("\u0600" <= c <= "\u06ff" for c in value):
                    value = format_arabic_text(value)
                    font_path = os.path.join(
                        PROJECT_FOLDER, "NotoNaskhArabic-Regular.ttf"
                    )
                    if os.path.exists(font_path):
                        new_page.insert_font("NotoNaskhArabic", font_path)
                        new_page.insert_text(
                            (rect.x0, rect.y1 - 2),
                            value,
                            fontsize=12,
                            fontname="NotoNaskhArabic",
                        )
                    else:
                        new_page.insert_text((rect.x0, rect.y1 - 2), value, fontsize=12)
                else:
                    font_path = os.path.join(PROJECT_FOLDER, "times.ttf")
                    if os.path.exists(font_path):
                        new_page.insert_font("Times", font_path)
                        new_page.insert_text(
                            (rect.x0, rect.y1 - 2), value, fontsize=12, fontname="Times"
                        )
                    else:
                        new_page.insert_text((rect.x0, rect.y1 - 2), value, fontsize=12)
        new_doc.save(output_path)
        new_doc.close()
        doc.close()
        logger.info(
            f"[REPLACE_DASHES] Finished replacing dashes and saved to {output_path}"
        )
        return True
    except Exception as e:
        logger.error(f"Error replacing dashes in PDF: {str(e)}")
        return False


# Add this mapping dictionary at the top of the file, after the imports
TEMPLATE_MAPPING = {
    "platinum": "platinum.pdf",
    "corporate": "Corporate.pdf",
    "business": "Business.pdf",
    "isic": "ISIC.pdf",
    "itic": "ITIC.pdf",
    "iytc": "IYTC.pdf",
    "cardcollection": "cardcollection.pdf",
}


def generate_pdf_for_card_type(
    row_data, output_path, card_type="a4", template_path=None
):
    """Generate a PDF by overlaying data on the template if provided, otherwise create a B5-size PDF from scratch for standard card types."""
    try:
        logger.info(f"Card type: {card_type}")

        # Get the appropriate template based on card type
        template_mapping = {
            "platinum": "platinum.pdf",
            "corporate": "Corporate.pdf",
            "business": "Business.pdf",
            "isic": "ISIC.pdf",
            "itic": "ITIC.pdf",
            "iytc": "IYTC.pdf",
            "a4": "cardcollection.pdf",
        }

        # If no template path provided, use the default template for the card type
        if not template_path and card_type in template_mapping:
            template_path = os.path.join(
                PROJECT_FOLDER, "static", "card_templates", template_mapping[card_type]
            )
            logger.info(f"Using default template for {card_type}: {template_path}")

        logger.info(f"Template path: {template_path}")
        logger.info(f"Output path: {output_path}")
        logger.info(f"Row data received: {row_data}")

        if template_path and os.path.exists(template_path):
            # Get values in the correct order based on card type
            values = []
            if card_type == "platinum":
                values = [
                    row_data.get("activation code", ""),
                    row_data.get("last four digits", ""),
                    row_data.get("name", ""),
                    row_data.get("phone number", ""),
                    row_data.get("delivery address", ""),
                    row_data.get("address title", ""),
                    row_data.get("address description", ""),
                ]
            elif card_type == "corporate":
                values = [
                    row_data.get("activation code", ""),
                    row_data.get("name", ""),
                    row_data.get("last four digits", ""),
                    row_data.get("phone number", ""),
                    row_data.get("delivery address", ""),
                ]
            elif card_type == "business":
                values = [
                    row_data.get("activation code", ""),
                    row_data.get("last four digits", ""),
                    row_data.get("name", ""),
                    row_data.get("onboarding name", ""),
                    row_data.get("phone number", ""),
                    row_data.get("delivery address", ""),
                    row_data.get("address title", ""),
                    row_data.get("address description", ""),
                ]
            elif card_type == "isic":
                values = [
                    row_data.get("last four digits", ""),
                    row_data.get("delivery address", ""),
                    row_data.get("phone number", ""),
                    row_data.get("name", ""),
                    row_data.get("university", ""),
                ]
            elif card_type == "itic":
                values = [
                    row_data.get("last four digits", ""),
                    row_data.get("delivery address", ""),
                    row_data.get("phone number", ""),
                    row_data.get("name", ""),
                    row_data.get("university", ""),
                ]
            elif card_type == "iytc":
                values = [
                    row_data.get("last four digits", ""),
                    row_data.get("delivery address", ""),
                    row_data.get("phone number", ""),
                    row_data.get("name", ""),
                ]
            elif card_type == "a4":
                values = [
                    row_data.get("name", ""),
                    row_data.get("phone number", ""),
                    row_data.get("last four digits", ""),
                    row_data.get("delivery address", ""),
                ]
                replace_dashes_in_pdf(template_path, output_path, values)
                return True

            # Replace dashes in template with values
            success = replace_dashes_in_pdf(template_path, output_path, values)
            if success:
                logger.info(f"Successfully generated PDF at {output_path}")
                return True
            else:
                logger.error("Failed to replace dashes in template")
                return False
        else:
            # If no template provided, use the old method
            logger.info("No template provided, using default PDF generation")
            # B5 size in points
            B5_SIZE = (498.9, 708.7)
            pdf = FPDF(unit="pt", format=B5_SIZE)
            pdf.add_page()
            font_path = os.path.join(PROJECT_FOLDER, "NotoNaskhArabic-Regular.ttf")
            pdf.add_font("NotoNaskhArabic", "", font_path, uni=True)
            times_font_path = os.path.join(PROJECT_FOLDER, "times.ttf")
            pdf.add_font("Times", "", times_font_path, uni=True)

            def print_field(val, x, y, size=14, force_times=False):
                if not force_times and is_arabic(val):
                    pdf.set_font("NotoNaskhArabic", "", size)
                    pdf.set_xy(x, y)
                    pdf.cell(0, 10, txt=format_arabic_text(val), ln=False)
                else:
                    pdf.set_font("Times", "", size)
                    pdf.set_xy(x, y)
                    pdf.cell(0, 10, txt=str(val), ln=False)

            name = row_data.get("name", row_data.get("cardholder name", ""))
            phone = row_data.get("phone number", row_data.get("card phone number", ""))
            last_digits = row_data.get(
                "last four digits", row_data.get("card last digits", "")
            )
            address = row_data.get(
                "delivery address", row_data.get("delivery location", "")
            )
            activation_code = row_data.get("activation code", "")
            address_title = row_data.get("address title", "")
            address_description = row_data.get("address description", "")
            onboarding_name = row_data.get("onboarding name", "")
            university = row_data.get("university", "")
            POSITIONS = {
                "platinum": [
                    (288, 375),  # activation code
                    (258, 406),  # last four digits
                    (183, 435),  # name
                    (245, 465),  # phone number
                    (230, 507),  # delivery address
                    (123, 568),  # address title
                    (123, 585),  # address description
                ],
                "corporate": [
                    (155, 375),  # activation code
                    (152, 407),  # name
                    (152, 470),  # last four digits
                    (152, 525),  # phone number
                    (152, 585),  # delivery address
                ],
                "business": [
                    (280, 375),  # activation code
                    (290, 405),  # last four digits
                    (210, 430),  # name
                    (185, 445),  # onboarding name
                    (240, 470),  # phone number
                    (230, 507),  # delivery address
                    (123, 568),  # address title
                    (123, 585),  # address description
                ],
                "isic": [
                    (150, 423),  # last four digits
                    (150, 463),  # delivery address
                    (150, 503),  # phone number
                    (150, 538),  # name
                    (150, 575),  # university
                ],
                "itic": [
                    (150, 423),  # last four digits
                    (150, 463),  # delivery address
                    (150, 503),  # phone number
                    (150, 538),  # name
                    (150, 575),  # university
                ],
                "iytc": [
                    (150, 463),  # last four digits
                    (150, 503),  # delivery address
                    (150, 538),  # phone number
                    (150, 575),  # name
                ],
                "a4": [
                    (195, 340),  # name
                    (195, 380),  # phone number
                    (195, 415),  # last four digits
                    (195, 450),  # delivery address
                ],
            }
            positions = POSITIONS.get(card_type, POSITIONS["platinum"])
            if card_type == "platinum":
                print_field(activation_code, *positions[0], force_times=True)
                print_field(last_digits, *positions[1], force_times=True)
                print_field(name, *positions[2], force_times=True)
                print_field(phone, *positions[3], force_times=True)
                print_field(address, *positions[4], force_times=True)
                (
                    pdf.set_font("NotoNaskhArabic", "", 8)
                    if is_arabic(address_title)
                    else pdf.set_font("Times", "", 8)
                )
                pdf.set_xy(*positions[5])
                pdf.cell(
                    0,
                    10,
                    txt=(
                        format_arabic_text(address_title)
                        if is_arabic(address_title)
                        else str(address_title)
                    ),
                    ln=False,
                )
                (
                    pdf.set_font("NotoNaskhArabic", "", 8)
                    if is_arabic(address_description)
                    else pdf.set_font("Times", "", 8)
                )
                pdf.set_xy(*positions[6])
                pdf.cell(
                    0,
                    10,
                    txt=(
                        format_arabic_text(address_description)
                        if is_arabic(address_description)
                        else str(address_description)
                    ),
                    ln=False,
                )
            elif card_type == "corporate":
                print_field(activation_code, *positions[0], force_times=True)
                print_field(name, *positions[1], force_times=True)
                print_field(last_digits, *positions[2], force_times=True)
                print_field(phone, *positions[3], force_times=True)
                print_field(address, *positions[4], force_times=True)
            elif card_type == "business":
                print_field(activation_code, *positions[0], force_times=True)
                print_field(last_digits, *positions[1], force_times=True)
                print_field(name, *positions[2], force_times=True)
                print_field(onboarding_name, *positions[3], force_times=True)
                print_field(phone, *positions[4], force_times=True)
                print_field(address, *positions[5], force_times=True)
                (
                    pdf.set_font("NotoNaskhArabic", "", 8)
                    if is_arabic(address_title)
                    else pdf.set_font("Times", "", 8)
                )
                pdf.set_xy(*positions[6])
                pdf.cell(
                    0,
                    10,
                    txt=(
                        format_arabic_text(address_title)
                        if is_arabic(address_title)
                        else str(address_title)
                    ),
                    ln=False,
                )
                (
                    pdf.set_font("NotoNaskhArabic", "", 8)
                    if is_arabic(address_description)
                    else pdf.set_font("Times", "", 8)
                )
                pdf.set_xy(*positions[7])
                pdf.cell(
                    0,
                    10,
                    txt=(
                        format_arabic_text(address_description)
                        if is_arabic(address_description)
                        else str(address_description)
                    ),
                    ln=False,
                )
            elif card_type == "isic":
                print_field(last_digits, *positions[0], force_times=True)
                print_field(address, *positions[1], force_times=True)
                print_field(phone, *positions[2], force_times=True)
                print_field(name, *positions[3], force_times=True)
                print_field(university, *positions[4], force_times=True)
            elif card_type == "itic":
                print_field(last_digits, *positions[0], force_times=True)
                print_field(address, *positions[1], force_times=True)
                print_field(phone, *positions[2], force_times=True)
                print_field(name, *positions[3], force_times=True)
                print_field(university, *positions[4], force_times=True)
            elif card_type == "iytc":
                print_field(last_digits, *positions[0], force_times=True)
                print_field(address, *positions[1], force_times=True)
                print_field(phone, *positions[2], force_times=True)
                print_field(name, *positions[3], force_times=True)
            elif card_type == "a4":
                print_field(name, *positions[0], size=20, force_times=True)
                print_field(phone, *positions[1], size=20, force_times=True)
                print_field(last_digits, *positions[2], size=20, force_times=True)
                print_field(address, *positions[3], size=20, force_times=True)
            else:
                # Default: just print all fields
                y = 100
                for k, v in row_data.items():
                    print_field(v, 50, y, force_times=True)
                    y += 20
            pdf.output(output_path)
            logger.info(f"Successfully generated B5-size PDF at {output_path}")
            return True
    except Exception as e:
        logger.error(f"Error generating PDF: {str(e)}")
        logger.exception("Full traceback:")
        return False


def process_excel_with_template(
    excel_path, session_id, card_type, pdf_template=None, zip_filename=None
):
    """Process Excel file with the appropriate template based on card type, using robust header mapping."""
    try:
        update_progress(session_id, 5)
        import openpyxl

        wb = openpyxl.load_workbook(excel_path)
        sheet = wb.active
        # Clean up empty rows and columns
        for row in range(sheet.max_row, 0, -1):
            if all(cell.value is None for cell in sheet[row]):
                sheet.delete_rows(row)
        for col in range(sheet.max_column, 0, -1):
            if all(cell.value is None for cell in sheet[get_column_letter(col)]):
                sheet.delete_cols(col)
        headers = []
        for header in sheet[1]:
            if header.value:
                headers.append(str(header.value).strip().lower())
            else:
                headers.append("")
        total_rows = max(sheet.max_row - 1, 1)
        output_folder = os.path.join(GENERATED_PDFS_FOLDER, session_id)
        os.makedirs(output_folder, exist_ok=True)
        pdf_paths = []
        for idx, row in enumerate(
            sheet.iter_rows(min_row=2, values_only=True), start=1
        ):
            if not any(cell for cell in row):
                continue
            # Robust row_dict creation
            row_dict = {
                headers[i].lower(): (
                    str(row[i]).strip() if i < len(row) and row[i] is not None else ""
                )
                for i in range(len(headers))
            }
            mapped_row = {}
            for k, v in row_dict.items():
                k_lower = k.lower()
                mapped_key = HEADER_TO_FIELD.get(k_lower, k_lower)
                mapped_row[mapped_key] = v
            name = mapped_row.get("name", "")
            if not name:
                name = mapped_row.get("cardholder name", "")
            last_digits = mapped_row.get("last four digits", "")
            if not last_digits:
                last_digits = mapped_row.get("card last digits", "")
            name_clean = secure_filename(name.replace(" ", "_"))
            pdf_filename = f"{name_clean.replace('_', ' ')} , {last_digits}.pdf"
            pdf_path = os.path.join(output_folder, pdf_filename)
            generate_pdf_for_card_type(mapped_row, pdf_path, card_type, pdf_template)
            pdf_paths.append(pdf_filename)
            if total_rows > 0:
                completion_ratio = float(idx) / float(total_rows)
                progress_percent = int(completion_ratio * 85)
                base_progress = 5
                progress_value = int(base_progress) + int(progress_percent)
                progress_final = min(max(0, progress_value), 99)
                update_progress(session_id, progress_final)
            time.sleep(0.05)
        if zip_filename is None:
            zip_filename = f"{session_id}_{card_type}.zip"
        zip_path = os.path.join(TEMP_FOLDER, zip_filename)
        with zipfile.ZipFile(zip_path, "w") as zipf:
            for pdf_path in pdf_paths:
                full_path = os.path.join(output_folder, pdf_path)
                if os.path.exists(full_path):
                    zipf.write(full_path, pdf_path)
        update_progress(session_id, 100)
        zip_info_path = os.path.join(TEMP_FOLDER, f"{session_id}_zip_info.txt")
        with open(zip_info_path, "w") as f:
            f.write(zip_path)
    except Exception as e:
        logger.error(f"Error in Excel processing: {str(e)}")
        update_progress(session_id, -1)


# =====================================
# PDF Editor
# =====================================

# =====================================
# PDF to Word Conversion
# =====================================

# =====================================
# Word to PDF Conversion
# =====================================

# =====================================
# Template Builder
# =====================================


# =====================================
# Utility Routes
# =====================================
@app.route("/view_pdf/<session_id>/<path:filename>")
def view_pdf(session_id, filename):
    """View PDF file"""
    # Check if file is from PDF folder
    pdf_path = os.path.join(GENERATED_PDFS_FOLDER, f"{session_id}_{filename}")

    if not os.path.exists(pdf_path):
        # Check if file is from generated PDFs folder
        pdf_path = os.path.join(GENERATED_PDFS_FOLDER, session_id, filename)

        if not os.path.exists(pdf_path):
            logger.error(f"PDF not found: {pdf_path}")
            return jsonify({"error": "PDF file not found"}), 404

    try:
        return send_file(pdf_path, mimetype="application/pdf")
    except Exception as e:
        logger.error(f"Error sending PDF: {str(e)}")
        return jsonify({"error": f"Error serving PDF: {str(e)}"}), 500


@app.route("/download_all_pdfs")
def download_all_pdfs():
    """Download all generated PDFs as zip"""
    session_id = session.get("session_id")
    if not session_id:
        flash("No session found", "warning")
        return redirect(url_for("index"))

    # Path to file containing PDF paths
    paths_file = os.path.join(TEMP_FOLDER, f"{session_id}_pdf_paths.txt")

    if not os.path.exists(paths_file):
        flash("No PDFs have been generated yet", "warning")
        return redirect(url_for("index"))

    try:
        # Read PDF paths
        with open(paths_file, "r") as f:
            pdf_paths = f.read().splitlines()

        # Create a BytesIO object
        memory_file = BytesIO()

        # Create zip file
        with zipfile.ZipFile(memory_file, "w") as zf:
            pdf_folder = os.path.join(GENERATED_PDFS_FOLDER, session_id)

            for pdf_path in pdf_paths:
                full_path = os.path.join(pdf_folder, pdf_path)
                if os.path.exists(full_path):
                    zf.write(full_path, pdf_path)

        # Reset file pointer
        memory_file.seek(0)

        # Create timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        # Return zip file
        return send_file(
            memory_file,
            mimetype="application/zip",
            as_attachment=True,
            download_name=f"bank_statements_{timestamp}.zip",
        )
    except Exception as e:
        logger.error(f"Error creating zip: {str(e)}")
        flash(f"Error creating zip file: {str(e)}", "danger")
        return redirect(url_for("index"))


@app.route("/download_converted_pdf")
def download_converted_pdf():
    """Download converted PDF file"""
    if "converted_pdf_file" not in session:
        flash("No converted PDF file available", "warning")
        return redirect(url_for("index"))

    pdf_path = session["converted_pdf_file"]
    filename = os.path.basename(pdf_path)

    return send_file(
        pdf_path, mimetype="application/pdf", as_attachment=True, download_name=filename
    )


@app.route("/show_progress/<job_type>")
def show_progress(job_type):
    """Show progress page for a job"""
    return render_template("progress.html", job_type=job_type)


def cleanup_uploads_folder():
    """Clean up temporary files in the uploads folder while preserving important data"""
    try:
        # Files to preserve (case-insensitive)
        preserve_files = {
            "NotoNaskhArabic-Regular.ttf",
            "times.ttf",
            "platinum.pdf",
            "Corporate.pdf",
            "Business.pdf",
            "ISIC.pdf",
            "ITIC.pdf",
            "IYTC.pdf",
            "cardcollection.pdf",
        }

        # Directories to preserve
        preserve_dirs = {
            os.path.join(PROJECT_FOLDER, "static", "card_templates"),
            os.path.join(PROJECT_FOLDER, "static", "assets"),
        }

        # Only clean specific temporary directories
        temp_dirs = [
            os.path.join(UPLOAD_FOLDER, "temp"),
            os.path.join(UPLOAD_FOLDER, "pdf"),
        ]

        logger.info("Starting cleanup of temporary files...")

        # Clean temporary directories
        for directory in temp_dirs:
            if os.path.exists(directory):
                logger.info(f"Cleaning directory: {directory}")
                for filename in os.listdir(directory):
                    file_path = os.path.join(directory, filename)
                    try:
                        # Skip if file should be preserved
                        if os.path.basename(file_path).lower() in {
                            f.lower() for f in preserve_files
                        }:
                            logger.info(f"Preserving file: {file_path}")
                            continue

                        # Skip if directory should be preserved
                        if any(
                            os.path.commonpath([file_path]) == os.path.commonpath([d])
                            for d in preserve_dirs
                        ):
                            logger.info(f"Preserving directory: {file_path}")
                            continue

                        # Delete temporary files and directories
                        if os.path.isfile(file_path):
                            os.unlink(file_path)
                            logger.info(f"Deleted temporary file: {file_path}")
                        elif os.path.isdir(file_path):
                            import shutil

                            shutil.rmtree(file_path)
                            logger.info(f"Deleted temporary directory: {file_path}")
                    except Exception as e:
                        logger.error(f"Error processing {file_path}: {str(e)}")

        logger.info("Cleanup of temporary files completed successfully")
    except Exception as e:
        logger.error(f"Error during cleanup: {str(e)}")


@app.route("/download/<filename>")
def download_file(filename):
    """Download a generated file (ZIP or PDF) and clean up temporary files afterward"""
    session_id = session.get("session_id")
    if not session_id:
        flash("Session expired. Please try again.", "warning")
        return redirect(url_for("index"))

    # For ZIP files
    if filename.endswith(".zip"):
        # Check if zip_filename is in session
        if "zip_filename" in session and session["zip_filename"] == filename:
            # File should be in TEMP_FOLDER
            file_path = os.path.join(TEMP_FOLDER, filename)

            # If file not found, try to get path from zip_info file
            if not os.path.exists(file_path):
                zip_info_path = os.path.join(TEMP_FOLDER, f"{session_id}_zip_info.txt")
                if os.path.exists(zip_info_path):
                    with open(zip_info_path, "r") as f:
                        file_path = f.read().strip()

            if os.path.exists(file_path):
                response = send_file(
                    file_path,
                    mimetype="application/zip",
                    as_attachment=True,
                    download_name=filename,
                )

                # Clean up only temporary files after sending
                @response.call_on_close
                def cleanup():
                    cleanup_uploads_folder()

                return response

    # For PDF files
    elif filename.endswith(".pdf"):
        # Check the GENERATED_PDFS_FOLDER for files matching session_id
        pdf_folder = os.path.join(GENERATED_PDFS_FOLDER, session_id)
        if os.path.exists(pdf_folder):
            # Look for the specific file
            for file in os.listdir(pdf_folder):
                if file == filename:
                    file_path = os.path.join(pdf_folder, file)
                    response = send_file(
                        file_path,
                        mimetype="application/pdf",
                        as_attachment=True,
                        download_name=filename,
                    )

                    # Clean up only temporary files after sending
                    @response.call_on_close
                    def cleanup():
                        cleanup_uploads_folder()

                    return response

    # If all checks fail
    flash("Requested file not found", "danger")
    return redirect(url_for("index"))


@app.route("/clear_session")
def clear_session():
    """Clear session data"""
    session_id = session.get("session_id")

    if session_id:
        # Clean up session files
        try:
            # Clear temporary files
            for folder in [GENERATED_PDFS_FOLDER, TEMP_FOLDER]:
                for filename in os.listdir(folder):
                    if filename.startswith(f"{session_id}_"):
                        os.remove(os.path.join(folder, filename))

            # Clear generated PDFs folder
            pdf_folder = os.path.join(GENERATED_PDFS_FOLDER, session_id)
            if os.path.exists(pdf_folder):
                import shutil

                shutil.rmtree(pdf_folder)

            # Remove from progress tracker
            if session_id in progress_tracker:
                del progress_tracker[session_id]
        except Exception as e:
            logger.error(f"Error cleaning up session: {str(e)}")

    # Clear session data
    session.clear()

    flash("Session cleared successfully", "success")
    return redirect(url_for("index"))


# =====================================
# PDF and Conversion Functions
# =====================================


@app.errorhandler(404)
def page_not_found(e):
    """Handle 404 errors"""
    return render_template("error.html", error="Page not found"), 404


@app.errorhandler(500)
def internal_server_error(e):
    """Handle 500 errors"""
    return render_template("error.html", error="Internal server error"), 500


@app.route("/card_collection_preview", methods=["GET", "POST"])
def card_collection_preview():
    if request.method == "POST":
        if "file" not in request.files:
            flash("Please upload an Excel file", "danger")
            return redirect(request.url)
        excel_file = request.files["file"]
        if excel_file.filename == "":
            flash("Please select an Excel file", "danger")
            return redirect(request.url)
        if excel_file and allowed_file(excel_file.filename, "excel"):
            try:
                session_id = session.get("session_id", str(uuid.uuid4()))
                session["session_id"] = session_id
                # Always use the static cardcollection.pdf template
                template_path = os.path.join(
                    PROJECT_FOLDER, "static", "card_templates", "cardcollection.pdf"
                )
                df = pd.read_excel(excel_file)
                data = df.to_dict("records")
                columns = df.columns.tolist()
                excel_path = os.path.join(UPLOAD_FOLDER, f"{session_id}_data.xlsx")
                excel_file.save(excel_path)
                session["excel_file"] = excel_path
                card_type = request.form.get("card_type", "a4")
                session["card_type"] = card_type
                if data:
                    preview_pdf_path = os.path.join(
                        GENERATED_PDFS_FOLDER, f"{session_id}_preview.pdf"
                    )
                    generate_pdf_for_card_type(
                        data[0],
                        preview_pdf_path,
                        card_type=card_type,
                        template_path=template_path,
                    )
                    session["preview_pdf"] = preview_pdf_path
                return render_template(
                    "card_collection_preview.html",
                    data=data,
                    columns=columns,
                    preview_pdf=f"/view_pdf/{session_id}/preview.pdf",
                    card_type=card_type,
                )
            except Exception as e:
                logger.error(f"Error processing files: {str(e)}")
                flash(f"Error processing files: {str(e)}", "danger")
                return redirect(request.url)
        else:
            flash(
                "Invalid file type. Please upload Excel (.xlsx, .xls) file", "warning"
            )
            return redirect(request.url)
    return render_template("card_collection_preview.html", data=None, columns=None)


@app.route("/generate_collection_pdfs", methods=["POST"])
def generate_collection_pdfs():
    """Generate PDFs from Excel data using the static cardcollection.pdf template, using platinum field order."""
    if "excel_file" not in session:
        flash("Please upload an Excel file first", "warning")
        return redirect(url_for("card_collection_preview"))
    try:
        session_id = session["session_id"]
        card_type = (
            "platinum"  # Always use platinum field order for static card collection
        )
        df = pd.read_excel(session["excel_file"])
        # Always use the static cardcollection.pdf template
        template_path = os.path.join(
            PROJECT_FOLDER, "static", "card_templates", "cardcollection.pdf"
        )
        output_folder = os.path.join(GENERATED_PDFS_FOLDER, session_id)
        os.makedirs(output_folder, exist_ok=True)
        pdf_paths = []
        total_rows = len(df)
        for i, row in df.iterrows():
            name = str(row.get("name", row.get("cardholder name", f"card_{i}")))
            last_digits = str(
                row.get("last four digits", row.get("card last digits", ""))
            )
            name_clean = secure_filename(name.replace(" ", "_"))
            pdf_filename = f"{name_clean} , {last_digits}.pdf"
            pdf_path = os.path.join(output_folder, pdf_filename)
            mapped_row = {k.lower(): v for k, v in row.to_dict().items()}
            # Use only the four fields for cardcollection, in this order
            field_order = [
                "last four digits",
                "delivery address",
                "phone number",
                "name",
            ]
            values = [mapped_row.get(k, "") for k in field_order]
            logger.info(
                f"[CARD_COLLECTION] Using values for dash replacement: {values}"
            )
            # Overlay data onto dashes in the static template
            replace_dashes_in_pdf(template_path, pdf_path, values)
            pdf_paths.append(pdf_filename)
            progress = int((i + 1) / total_rows * 100)
            update_progress(session_id, progress)
        zip_filename = f"card_collection_{session_id}.zip"
        zip_path = os.path.join(TEMP_FOLDER, zip_filename)
        with zipfile.ZipFile(zip_path, "w") as zipf:
            for pdf_path in pdf_paths:
                full_path = os.path.join(output_folder, pdf_path)
                if os.path.exists(full_path):
                    zipf.write(full_path, pdf_path)
        session["zip_filename"] = zip_filename
        update_progress(session_id, 100)
        return jsonify(
            {
                "success": True,
                "message": "PDFs generated successfully",
                "download_url": url_for("download_file", filename=zip_filename),
            }
        )
    except Exception as e:
        logger.error(f"Error generating PDFs: {str(e)}")
        update_progress(session_id, -1)
        return (
            jsonify({"success": False, "message": f"Error generating PDFs: {str(e)}"}),
            500,
        )


# =====================================
# Simple Excel Table to PDF Conversion
# =====================================
@app.route("/simple_excel_to_pdf", methods=["GET", "POST"])
def simple_excel_to_pdf():
    """Convert Excel file to PDF with table (no mapping, just as-is)"""
    if request.method == "POST":
        if "file" not in request.files:
            flash("No file selected", "danger")
            return redirect(request.url)
        file = request.files["file"]
        if file.filename == "":
            flash("No file selected", "danger")
            return redirect(request.url)
        if file and allowed_file(file.filename, "excel"):
            session_id = session.get("session_id", str(uuid.uuid4()))
            session["session_id"] = session_id
            filename = secure_filename(file.filename)
            excel_path = os.path.join(UPLOAD_FOLDER, f"{session_id}_simple_{filename}")
            file.save(excel_path)
            try:
                pdf_filename = os.path.splitext(filename)[0] + ".pdf"
                pdf_path = os.path.join(
                    GENERATED_PDFS_FOLDER, f"{session_id}_{pdf_filename}"
                )
                excel_table_to_pdf(excel_path, pdf_path)
                session["simple_pdf_file"] = pdf_path
                session["simple_pdf_name"] = pdf_filename
                flash("Excel table successfully converted to PDF", "success")
                return redirect(url_for("download_simple_pdf"))
            except Exception as e:
                logger.error(f"Error converting Excel to PDF: {str(e)}")
                flash(f"Error converting Excel to PDF: {str(e)}", "danger")
                return redirect(request.url)
        else:
            flash("Invalid file type. Allowed types: xlsx, xls", "warning")
            return redirect(request.url)
    return render_template("simple_excel_to_pdf.html")


@app.route("/download_simple_pdf")
def download_simple_pdf():
    """Download the simple Excel table PDF"""
    if "simple_pdf_file" not in session:
        flash("No converted PDF file available", "warning")
        return redirect(url_for("simple_excel_to_pdf"))
    pdf_path = session["simple_pdf_file"]
    filename = session.get("simple_pdf_name", os.path.basename(pdf_path))
    return send_file(
        pdf_path, mimetype="application/pdf", as_attachment=True, download_name=filename
    )


def kurdish_to_arabic_chars(text):
    """Replace Kurdish-specific chars with Arabic equivalents. For 'ە' inside a word, change to Arabic 'ه' and add a space if not followed by a space. If 'ە' is a separate word or followed by a space, keep it as is."""
    if not isinstance(text, str):
        return text
    # Kurdish 'ێ' (U+06CE) -> Arabic 'ي' (U+064A)
    text = text.replace("\u06ce", "\u064a")
    import re

    # Replace 'ە' inside a word (not surrounded by spaces)
    def replace_inside_word(match):
        next_char = match.group(2)
        if next_char == " ":
            return "\u0647 "
        else:
            return "\u0647 "

    # Only replace 'ە' that is not already followed by a space
    text = re.sub(r"(?<! )\u06d5([^ ])", lambda m: "\u0647 " + m.group(1), text)
    return text


def excel_table_to_pdf(excel_path, pdf_path):
    """Convert an Excel file to a PDF with a table (all rows/columns), supporting Kurdish, Arabic, and English."""
    import pandas as pd
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Spacer
    from reportlab.lib import colors
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.pdfbase import pdfmetrics
    import os

    df = pd.read_excel(excel_path)
    PROJECT_FOLDER = os.path.dirname(os.path.abspath(__file__))
    font_path = os.path.join(PROJECT_FOLDER, "NotoNaskhArabic-Regular.ttf")
    font_name = "NotoNaskhArabic"
    if os.path.exists(font_path):
        pdfmetrics.registerFont(TTFont(font_name, font_path))
    else:
        font_name = "Helvetica"  # fallback

    def format_cell(val):
        try:
            from bidi.algorithm import get_display
            from arabic_reshaper import reshape
        except ImportError:
            return str(val)
        if val is None:
            return ""
        s = str(val)
        s = kurdish_to_arabic_chars(s)
        if any("\u0600" <= c <= "\u06ff" for c in s):
            try:
                reshaped = reshape(s)
                return get_display(reshaped)
            except Exception:
                return s
        return s

    data = [[format_cell(col) for col in df.columns.tolist()]]
    for row in df.values.tolist():
        data.append([format_cell(cell) for cell in row])
    doc = SimpleDocTemplate(pdf_path, pagesize=A4)
    elements = []
    table = Table(data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("FONTNAME", (0, 0), (-1, -1), font_name),
                ("FONTSIZE", (0, 0), (-1, 0), 12),
                ("FONTSIZE", (0, 1), (-1, -1), 10),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 10),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )
    elements.append(table)
    elements.append(Spacer(1, 12))
    doc.build(elements)


if __name__ == "__main__":
    # Create required directories if they don't exist
    for folder in [UPLOAD_FOLDER, TEMPLATES_FOLDER, GENERATED_PDFS_FOLDER, TEMP_FOLDER]:
        os.makedirs(folder, exist_ok=True)

    # Get configuration from environment variables
    host = os.environ.get("HOST", "0.0.0.0")
    port = int(os.environ.get("PORT", 5656))
    debug = os.environ.get("FLASK_DEBUG", "0") == "1"

    try:
        app.run(host=host, port=port, debug=debug)
    finally:
        # Shutdown the scheduler when the app stops
        scheduler.shutdown()